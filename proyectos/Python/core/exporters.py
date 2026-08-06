"""
core.exporters
===============
Funciones de exportación reutilizables por cualquier módulo de la
suite: CSV, Excel (.xlsx) y PDF. Reciben datos genéricos (lista de
diccionarios) para no acoplarse a ningún dominio en particular.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Error controlado durante una operación de exportación."""


def _validar_datos(datos: list[dict[str, Any]]) -> None:
    if not datos:
        raise ExportError("No hay datos para exportar.")


def timestamped_filename(base_name: str, extension: str) -> str:
    """Genera un nombre de archivo único con marca de tiempo."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{ts}.{extension}"


# CSV
def export_to_csv(datos: list[dict[str, Any]], destino: Path) -> Path:
    _validar_datos(datos)
    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        with open(destino, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(datos[0].keys()))
            writer.writeheader()
            writer.writerows(datos)
        logger.info("Exportado a CSV: %s", destino)
        return destino
    except OSError as exc:
        raise ExportError(f"No se pudo escribir el CSV: {exc}") from exc


# Excel (openpyxl)
def export_to_excel(
    datos: list[dict[str, Any]], destino: Path, titulo_hoja: str = "Datos"
) -> Path:
    _validar_datos(datos)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ExportError(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl"
        ) from exc

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = titulo_hoja[:31]  # límite de Excel para nombres de hoja

        encabezados = list(datos[0].keys())
        header_fill = PatternFill(start_color="1F6AA5", end_color="1F6AA5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, fila in enumerate(datos, start=2):
            for col_idx, encabezado in enumerate(encabezados, start=1):
                ws.cell(row=row_idx, column=col_idx, value=fila.get(encabezado))

        # Autoajuste aproximado de ancho de columnas
        for col_idx, encabezado in enumerate(encabezados, start=1):
            max_len = max(
                [len(str(encabezado))] + [len(str(fila.get(encabezado, ""))) for fila in datos]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
        logger.info("Exportado a Excel: %s", destino)
        return destino
    except OSError as exc:
        raise ExportError(f"No se pudo escribir el Excel: {exc}") from exc


# PDF (reportlab)
def export_to_pdf(
    datos: list[dict[str, Any]], destino: Path, titulo: str = "Reporte"
) -> Path:
    _validar_datos(datos)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
    except ImportError as exc:
        raise ExportError(
            "reportlab no está instalado. Ejecuta: pip install reportlab"
        ) from exc

    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            str(destino),
            pagesize=letter,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )
        estilos = getSampleStyleSheet()
        elementos = [
            Paragraph(titulo, estilos["Title"]),
            Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                estilos["Normal"],
            ),
            Spacer(1, 0.5 * cm),
        ]

        encabezados = list(datos[0].keys())
        filas_tabla = [[h.replace("_", " ").title() for h in encabezados]]
        for fila in datos:
            filas_tabla.append([str(fila.get(h, "")) for h in encabezados])

        tabla = Table(filas_tabla, repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6AA5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ]
            )
        )
        elementos.append(tabla)
        doc.build(elementos)
        logger.info("Exportado a PDF: %s", destino)
        return destino
    except OSError as exc:
        raise ExportError(f"No se pudo escribir el PDF: {exc}") from exc
